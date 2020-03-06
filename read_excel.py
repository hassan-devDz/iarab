from openpyxl import load_workbook
import re

filename="/home/hassan/empty_book1.xlsx"
workbook = load_workbook(filename=filename)
sheet = workbook.active
COMMA            = u'\u060C'
SEMICOLON        = u'\u061B'
QUESTION         = u'\u061F'
HAMZA            = u'\u0621'
ALEF_MADDA       = u'\u0622'
ALEF_HAMZA_ABOVE = u'\u0623'
WAW_HAMZA        = u'\u0624'
ALEF_HAMZA_BELOW = u'\u0625'
YEH_HAMZA        = u'\u0626'
ALEF             = u'\u0627'
BEH              = u'\u0628'
TEH_MARBUTA      = u'\u0629'
TEH              = u'\u062a'
THEH             = u'\u062b'
JEEM             = u'\u062c'
HAH              = u'\u062d'
KHAH             = u'\u062e'
DAL              = u'\u062f'
THAL             = u'\u0630'
REH              = u'\u0631'
ZAIN             = u'\u0632'
SEEN             = u'\u0633'
SHEEN            = u'\u0634'
SAD              = u'\u0635'
DAD              = u'\u0636'
TAH              = u'\u0637'
ZAH              = u'\u0638'
AIN              = u'\u0639'
GHAIN            = u'\u063a'
TATWEEL          = u'\u0640'
FEH              = u'\u0641'
QAF              = u'\u0642'
KAF              = u'\u0643'
LAM              = u'\u0644'
MEEM             = u'\u0645'
NOON             = u'\u0646'
HEH              = u'\u0647'
WAW              = u'\u0648'
ALEF_MAKSURA     = u'\u0649'
YEH              = u'\u064a'
MADDA_ABOVE      = u'\u0653'
HAMZA_ABOVE      = u'\u0654'
HAMZA_BELOW      = u'\u0655'
ZERO             = u'\u0660'
ONE              = u'\u0661'
TWO              = u'\u0662'
THREE            = u'\u0663'
FOUR             = u'\u0664'
FIVE             = u'\u0665'
SIX              = u'\u0666'
SEVEN            = u'\u0667'
EIGHT            = u'\u0668'
NINE             = u'\u0669'
PERCENT          = u'\u066a'
DECIMAL          = u'\u066b'
THOUSANDS        = u'\u066c'
STAR             = u'\u066d'
MINI_ALEF        = u'\u0670'
ALEF_WASLA       = u'\u0671'
FULL_STOP        = u'\u06d4'
BYTE_ORDER_MARK  = u'\ufeff'

# Diacritics
FATHATAN         = u'\u064b'
DAMMATAN         = u'\u064c'
KASRATAN         = u'\u064d'
FATHA            = u'\u064e'
DAMMA            = u'\u064f'
KASRA            = u'\u0650'
SHADDA           = u'\u0651'
SUKUN            = u'\u0652'

#Ligatures
LAM_ALEF=u'\ufefb'
LAM_ALEF_HAMZA_ABOVE=u'\ufef7'
LAM_ALEF_HAMZA_BELOW=u'\ufef9'
LAM_ALEF_MADDA_ABOVE=u'\ufef5'
simple_LAM_ALEF=u'\u0644\u0627'
simple_LAM_ALEF_HAMZA_ABOVE=u'\u0644\u0623'
simple_LAM_ALEF_HAMZA_BELOW=u'\u0644\u0625'
simple_LAM_ALEF_MADDA_ABOVE=u'\u0644\u0622'
def fatha(w):
    lst=[]
    for x in w :
        lst.append(x+FATHA)
    return ''.join(lst)
def fiaal_moatal_yeh(elfiaal):
    tsahih= elfiaal[0:4]
    return tsahih+ALEF_MAKSURA
def fiaal_moatal_waw(elfiaal):
    tsahih= elfiaal[0:4]
    return tsahih+ALEF
def men_faala_il_faaila(elfiaal):
    tashih = elfiaal[0:3]+KASRA+elfiaal[4:6]
    return tashih
ls=[] #قائمة الافعال المعالجة
for value in sheet.iter_rows(min_row=2,max_row=6587,min_col=15,max_col=20,values_only=True):

    string2 = re.sub ( r"\s+" , "" , str(value [ 4 ]) )

    jader = fatha(string2) #الجذور مشكلة بالفتحة
    print(jader)

    if HAMZA in jader and value[0] != None:

        fiaal = jader.replace ( HAMZA , ALEF_HAMZA_ABOVE ) #قلب الهمزة الى الف و همزة في الجذور لتصبح أفعال
        print('THAT G' , fiaal )
        if fiaal[2 ]==fiaal[4 ]: #معالجة الحرف المضعف في الفعل
            L= fiaal.replace( fiaal[ 2:6 ] , fiaal[2 ] + SHADDA + FATHA )
            #print(L,value[5])
            ls.append(L)
            #sheet[e.value]=L
            #print(sheet['T7'])
            #sheet['T1':'T10']=L
        elif jader [ 4 ] == YEH: #معالجة الفعل المعتل الناقص اليائي
            y= fiaal_moatal_yeh(fiaal)
            ls.append(y)
        elif jader [ 4 ] == WAW: #معالجة الفعل المعتل الناقص اليائي
            W= fiaal_moatal_waw(fiaal)
            ls.append(W)

        else:
            ls.append( fiaal )
            #print (sheet['T7'])
            #print ("G = {}".format (G))
    elif HAMZA not in jader and value [ 0 ] != None:
        print()
    else:
        ls.append(value[0])

print(ls)
def ktaba_fi_excel(start_row=int,column=int,lst=list):
    for i in range ( 0 , len ( lst ) ):
        e = sheet.cell ( row=i + start_row , column=column )
        e.value = ls [ i ]

#start_row = sheet.max_column - 3

ktaba_fi_excel(2,20,ls)
workbook.save(filename)
"""
wb = Workbook()
ws = wb.active
a1 = ws['A1']
ft = Font(color=colors.RED,name='KacstOne', size=16)
a1.font=ft
ws['A1']='lololo'
wb.save(filename)
"""
workbook.close()

